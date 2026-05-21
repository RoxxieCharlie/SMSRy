from __future__ import annotations

import csv
from io import BytesIO
from datetime import datetime, timedelta, time

from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import render
from django.utils import timezone

from store.models import IssuanceItem


# New-report cutover time: Sunday 6:00 PM
SUNDAY_EVENING_START_HOUR = 18  # 6 PM


def _style_weekly_sheet(ws, title, subtitle, headers, rows, totals=None):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    dark = "0F172A"
    blue = "1D4ED8"
    muted = "64748B"
    soft = "EAF2FF"
    line = "CBD5E1"

    ws.sheet_view.showGridLines = False

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"] = title.upper()
    ws["A1"].font = Font(size=20, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=dark)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws["A2"] = subtitle
    ws["A2"].font = Font(size=12, bold=True, color=dark)
    ws["A2"].fill = PatternFill("solid", fgColor=soft)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 24

    current_row = 4
    if totals:
        for idx, (label, value) in enumerate(totals, start=1):
            cell = ws.cell(row=current_row, column=idx)
            cell.value = f"{label}: {value}"
            cell.font = Font(bold=True, color=dark)
            cell.fill = PatternFill("solid", fgColor=soft)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(bottom=Side(style="thin", color=line))
        current_row += 2

    header_row = current_row
    ws.freeze_panes = f"A{header_row + 1}"
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="thin", color="93C5FD"))

    for row_index, row in enumerate(rows, start=header_row + 1):
        fill = PatternFill("solid", fgColor="F8FAFC" if row_index % 2 == 0 else "FFFFFF")
        for col, value in enumerate(row, start=1):
            cell = ws.cell(row=row_index, column=col)
            cell.value = value
            cell.fill = fill
            header = headers[col - 1].lower()
            is_centered = header in {"s/n", "total issued", "total collected", "share"}
            cell.alignment = Alignment(
                horizontal="center" if is_centered else "left",
                vertical="top",
                wrap_text=True,
            )
            cell.border = Border(bottom=Side(style="thin", color="E2E8F0"))

    last_row = max(header_row + len(rows), header_row)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{last_row}"

    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = len(str(headers[col - 1]))
        for row in range(header_row + 1, last_row + 1):
            value = ws.cell(row=row, column=col).value
            max_len = max(max_len, len(str(value or "")))
        ws.column_dimensions[letter].width = min(max(max_len + 4, 14), 44)


def _build_weekly_excel(report_list, dept_summary, weekly_total, start_dt, end_dt, generated_at):
    from openpyxl import Workbook

    wb = Workbook()
    summary = wb.active
    summary.title = "Weekly Report"

    period = f"{start_dt.date()} to {end_dt.date()}"
    generated = generated_at.strftime("%d %b %Y %H:%M")

    item_rows = [
        [idx, row["item"], row["total_quantity"], row["departments_with_usage"]]
        for idx, row in enumerate(report_list, start=1)
    ]
    _style_weekly_sheet(
        summary,
        "SMS WEEKLY REPORT",
        f"REPORT PERIOD: {period} | GENERATED: {generated}",
        ["S/N", "Item", "Total Issued", "Departments with Usage"],
        item_rows,
        totals=[
            ("Total issued", weekly_total),
            ("Unique items", len(report_list)),
            ("Departments", len(dept_summary)),
        ],
    )

    departments = wb.create_sheet("Department Summary")
    dept_rows = [
        [
            idx,
            row["department"],
            row["total_quantity"],
            f'{row["pct"]}%',
            ", ".join(f'{i["item"]} ({i["qty"]})' for i in row.get("items", [])),
        ]
        for idx, row in enumerate(dept_summary, start=1)
    ]
    _style_weekly_sheet(
        departments,
        "SMS WEEKLY DEPARTMENT SUMMARY",
        f"REPORT PERIOD: {period} | GENERATED: {generated}",
        ["S/N", "Department", "Total Collected", "Share", "Items Collected"],
        dept_rows,
        totals=[
            ("Total issued", weekly_total),
            ("Departments", len(dept_summary)),
        ],
    )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ------------------------------------------------------------
# OLD TIME-BOUND ACCESS WINDOW (COMMENTED OUT FOR NOW)
# Keep it here so you can re-enable later if you truly want.
# ------------------------------------------------------------
# def _weekly_report_window_open(now):
#     weekday = now.weekday()  # Mon=0 ... Sun=6
#     hour = now.hour
#     if weekday == 6:
#         return (hour >= SUNDAY_EVENING_START_HOUR, "Sunday 6:00 PM → Tuesday 11:59 PM")
#     if weekday in (0, 1):
#         return (True, "Sunday 6:00 PM → Tuesday 11:59 PM")
#     return (False, "Sunday 6:00 PM → Tuesday 11:59 PM")


def _week_bounds_from_monday(monday_date, tzinfo):
    """
    monday_date: date object for Monday
    returns aware datetimes:
      start: Monday 00:00:00
      end:   Sunday 23:59:59.999999
    """
    start_dt = timezone.make_aware(datetime.combine(monday_date, time.min), tzinfo)
    sunday_date = monday_date + timedelta(days=6)
    end_dt = timezone.make_aware(datetime.combine(sunday_date, time.max), tzinfo)
    return start_dt, end_dt


def _report_week_range(now_local):
    """
    Rule you asked for:
      - Reports are Monday → Sunday (full week)
      - A new one becomes available ONLY on Sunday 6PM.
      - Until Sunday 6PM, users keep seeing the PREVIOUS completed week.
      - From Sunday 6PM onward, users see the CURRENT week (Mon→Sun) as the "just concluded" week.

    Returns: (start_dt, end_dt, window_label, is_new_week_available)
    """
    tzinfo = now_local.tzinfo

    # Monday of the current week
    this_monday_date = (now_local - timedelta(days=now_local.weekday())).date()

    is_sunday = (now_local.weekday() == 6)
    is_after_cutoff = is_sunday and (now_local.hour >= SUNDAY_EVENING_START_HOUR)

    if is_after_cutoff:
        # New report is available: show current week Mon→Sun
        start_dt, end_dt = _week_bounds_from_monday(this_monday_date, tzinfo)
        window_label = f"{start_dt.date()} → {end_dt.date()}"
        return start_dt, end_dt, window_label, True

    # Not yet Sunday 6PM: show previous completed week Mon→Sun
    prev_monday_date = this_monday_date - timedelta(days=7)
    start_dt, end_dt = _week_bounds_from_monday(prev_monday_date, tzinfo)
    window_label = f"{start_dt.date()} → {end_dt.date()}"
    return start_dt, end_dt, window_label, False


@login_required
def weekly_report(request):
    user = request.user

    # Management only (case-insensitive safer)
    if not user.groups.filter(name__iexact="Management").exists():
        return HttpResponseForbidden("You do not have access to this page.")

    now = timezone.localtime(timezone.now())

    # ✅ Correct weekly range per your spec
    start_dt, end_dt, window_label, new_week_available = _report_week_range(now)

    # Only count valid (non-reversed) issuances
    qs = (
        IssuanceItem.objects
        .filter(
            issuance__is_reversed=False,
            issuance__issued_at__gte=start_dt,
            issuance__issued_at__lte=end_dt,
        )
        .select_related("item", "issuance__staff__department")
    )

    # Weekly total (all quantities issued)
    weekly_total = qs.aggregate(total=Sum("quantity"))["total"] or 0

    # Aggregate:
    # item_name -> {"total": int, "dept_counts": {dept_name: int}}
    agg = {}
    dept_totals = {}  # dept_name -> total_qty
    dept_items = {}   # dept_name -> {item_name: qty}

    for line in qs:
        item_name = line.item.name
        dept_name = (
            line.issuance.staff.department.name
            if line.issuance.staff.department
            else "Unassigned"
        )

        if item_name not in agg:
            agg[item_name] = {"total": 0, "dept_counts": {}}

        agg[item_name]["total"] += line.quantity
        agg[item_name]["dept_counts"][dept_name] = (
            agg[item_name]["dept_counts"].get(dept_name, 0) + line.quantity
        )

        dept_totals[dept_name] = dept_totals.get(dept_name, 0) + line.quantity

        if dept_name not in dept_items:
            dept_items[dept_name] = {}
        dept_items[dept_name][item_name] = dept_items[dept_name].get(item_name, 0) + line.quantity

    # Build item rows
    report_list = []
    for item_name, data in agg.items():
        dept_parts = [
            f"{dept} ({qty})"
            for dept, qty in sorted(data["dept_counts"].items(), key=lambda x: (-x[1], x[0]))
        ]
        report_list.append({
            "item": item_name,
            "total_quantity": data["total"],
            "departments_with_usage": ", ".join(dept_parts),
        })

    report_list.sort(key=lambda r: (-r["total_quantity"], r["item"]))

    # Build department summary rows
    dept_summary = []
    for dept_name, qty in sorted(dept_totals.items(), key=lambda x: (-x[1], x[0])):
        pct = 0 if weekly_total == 0 else round((qty / weekly_total) * 100)
        items_for_dept = sorted(
            dept_items.get(dept_name, {}).items(),
            key=lambda x: (-x[1], x[0]),
        )
        dept_summary.append({
            "department": dept_name,
            "total_quantity": qty,
            "pct": pct,
            "items": [{"item": n, "qty": q} for n, q in items_for_dept],
        })

    # Export a styled Excel workbook for management reporting.
    if request.GET.get("export") == "1":
        try:
            workbook = _build_weekly_excel(
                report_list,
                dept_summary,
                weekly_total,
                start_dt,
                end_dt,
                now,
            )
        except ImportError:
            response = HttpResponse(content_type="text/csv")
            filename = f"weekly_report_{start_dt.date()}_to_{end_dt.date()}.csv"
            response["Content-Disposition"] = f'attachment; filename="{filename}"'

            writer = csv.writer(response)
            writer.writerow(["Item", "Total Issued", "Departments with Usage"])
            for row in report_list:
                writer.writerow([row["item"], row["total_quantity"], row["departments_with_usage"]])
            return response

        response = HttpResponse(
            workbook.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        filename = f"weekly_report_{start_dt.date()}_to_{end_dt.date()}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    return render(request, "store/weekly_v2.html", {
        "active_nav": "weekly",

        # always accessible
        "report_closed": False,

        # period
        "start_date": start_dt.date(),
        "end_date": end_dt.date(),
        "window_label": window_label,

        # whether current week's report has switched in (Sun 6PM+)
        "new_week_available": new_week_available,
        "next_cutover_hint": "New report becomes available Sunday 6:00 PM.",

        # totals
        "weekly_total": weekly_total,

        # tables
        "report_list": report_list,
        "dept_summary": dept_summary,
    })
